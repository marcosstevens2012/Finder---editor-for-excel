#!/usr/bin/env python
#-*- coding: utf-8 -*-
# -*- coding: cp1252 -*-
import sys
reload(sys)
sys.setdefaultencoding('utf8')
import sys,os
from PyQt4.QtCore import *
from decimal import Decimal
from PyQt4 import QtGui, uic
import re
from PyQt4 import QtCore, QtGui, uic
from AcercaDe import acercade
from openpyxl import*



wb = Workbook()
# grab the active worksheet
ws = wb.active
import openpyxl   #Importamos libreria para abrir xls


##file_location = 'C:\Users\Richard\Dropbox\PROYECTO BUSCADOR/address.xlsx' #declaramos path o ruta del archivo a abrir

#se importa la pantalla ""from (carpeta) import (archivo)""

form_class = uic.loadUiType("Buscador.ui")[0]
###################################################################################

class Buscador(QtGui.QMainWindow, form_class):
    def __init__(self, parent=None):
        QtGui.QMainWindow.__init__(self, parent)
        self.setupUi(self)
        #self.l_cod_b.setValidator(QtGui.QIntValidator())
        self.l_carga.hide() == True ##oculta el label carga
        self.btn_cargar.setEnabled(False) 
        self.array_2_table()
        self.btn_cargar.clicked.connect(self.btn_cargar_clicked)##metodo carga archivo
        #self.btn_buscar.clicked.connect(self.btn_Buscar_clicked)
        self.btn_AcercaDe.clicked.connect(self.btn_AcercaDe_clicked)
        self.l_porc_b.textChanged.connect(self.l_porc_b_Changed)
        self.tabla.clicked.connect(self.mostrarDatos)
        self.l_cod.textChanged.connect(self.buscar_codigo)
        self.l_cod_b.setEnabled(False)
        self.l_cod.setEnabled(False)
        self.l_nombre.setEnabled(False)
        self.btn_buscar.setEnabled(False)
        self.l_porc_b.setFocus()
        self.list_precio_iva_orig=[]#lista con los precios son iva original
        self.list_precio_iva_new=[]#lista con los precios iva con ganancia
        self.list_codigo=[]#carga los codigo
        self.list_codigo_scanner=[]
        self.list_descripcion=[]

######################################################################################################################


    def array_2_table(self):
        self.tabla.setRowCount(18000)
        QtGui.QMessageBox.information( self,'Informacion', '''Por favor defina porcentaje de ganancia antes de cargar el archivo''', QtGui.QMessageBox.Ok)



######################################################################################################################
    def buscar_codigo(self):
        while self.tabla.rowCount() > 0:  # limpia la tabla de pacientes
            self.tabla.removeRow(0)

        self.codigo_b = str(self.l_cod.text())
        #Se agregan los elementos al QListWidget
        self.lon=len(self.l_cod.text())
        print str(self.lon)
        i = 0
        y=-1
        for x in self.list_codigo:
            self.cod=str(self.list_codigo)[0:self.lon]
            y+=1
            if (self.codigo_b)== str(self.cod):
                self.tabla.setItem(i, 0, QtGui.QTableWidgetItem(self.list_codigo_scanner[y]))
                self.tabla.setItem(i, 1, QtGui.QTableWidgetItem(self.list_descripcion[y]))
                self.tabla.setItem(i, 2, QtGui.QTableWidgetItem("{0:.2f}".format(self.list_precio_iva_new[y])))#se redondeo el valor a dos decimales
                self.tabla.setItem(i, 3, QtGui.QTableWidgetItem(str(self.list_codigo[y])))
                #self.tabla.setEditTriggers(QtGui.QAbstractItemView.NoEditTriggers)#metodo para q no se puedan editar las columnas
                #self.tabla.setSelectionBehavior(self.tabla.SelectRows)
                i += 1


#########################################################################################################
    def mostrarDatos(self):#muestra los datos en los labels
        self.tabla.setSelectionBehavior(self.tabla.SelectRows)
        rows = self.tabla.selectionModel().selectedRows()
        index = []
        for i in rows:
            index.append(i.row())
        for i in index:
            self.descripcion = self.tabla.item(i, 1).text()
            self.precio=self.tabla.item(i, 2).text()
            #self.nomGrupo = self.tableGrupos.item(i, 2).text()
            #self.estado = self.tableGrupos.item(i, 1).text()
            self.l_descripcion.setText(str(self.descripcion))
            self.l_precio_iva.setText(str(self.precio))
            self.precio_orig_iva=("{0:.2f}".format(self.list_precio_iva_orig[i]))
            self.l_ganancia.setText(str(float(self.precio)-float(self.precio_orig_iva)))


    #if (l_porc_b.isModified):
    def l_porc_b_Changed(self): 
  		self.btn_cargar.setEnabled(True)


    def btn_AcercaDe_clicked(self):
        from AcercaDe import acercade
     	ventana_2 = acercade(self)
        #se muestra la ventana
     	ventana_2.show()


#   #######################################################################
    def btn_cargar_clicked(self):#abre el explorador de archivos
    	self.filePath = QtGui.QFileDialog.getOpenFileName(self,'Cargar Archivo',"~/Desktop/",'*.xlsx')
        #doc = openpyxl.load_workbook(str(self.filePath))
        #hoja = doc.get_sheet_by_name('Sheet1')  # abre el archivo #se debe poner en el metodo a usar
        self.l_carga.show()#muestra el label carga
        self.l_carga.setStyleSheet("QLabel#l_carga {color: green}")#cambia de color el label carga
        self.l_cod_b.setEnabled(True)
        self.l_cod.setEnabled(True)
        self.l_nombre.setEnabled(True)
        self.btn_buscar.setEnabled(True)
        print(str(self.filePath))
        doc = openpyxl.load_workbook(str(self.filePath))
        hoja = doc.get_sheet_by_name('Hoja1')


    	seleccionCodigo = hoja['A18':'A17833'] 
        seleccionDesc = hoja['B18':'B17833']
        seleccionScanner = hoja['F18':'F17833']
        seleccionIVA = hoja['K18':'K17833']


        porcentaje = self.l_porc_b.text()
        porcentaje2 = (float(porcentaje) / 100)
        p = 0.0+porcentaje2

        f_i = 0

        for filas in seleccionCodigo:
            for celda in filas:
                self.tabla.setItem(f_i,3, QtGui.QTableWidgetItem(str(celda.value)))
                self.list_codigo.append(celda.value)
            f_i +=1

        f_i = 0
        for filas in seleccionDesc:
                for celda in filas:
                        self.tabla.setItem(f_i,1, QtGui.QTableWidgetItem(celda.value))
                        self.list_descripcion.append(celda.value)
                f_i +=1


        f_i = 0
        for filas in seleccionScanner:
                for celda in filas:
                    self.tabla.setItem(f_i,0, QtGui.QTableWidgetItem(str(celda.value)))
                    self.list_codigo_scanner.append(celda.value)
                f_i +=1

        f_i = 0
        p1 = 0
        self.i=0
        for filas in seleccionIVA:
            for celda in filas:

                p1 = celda.value*porcentaje2
                self.tabla.setItem(f_i,2, QtGui.QTableWidgetItem("{0:.2f}".format(p1+celda.value)))
                self.list_precio_iva_new.append(p1+celda.value)#agrega datos a la lista en la ulrima posicion
                self.list_precio_iva_orig.append(celda.value)
                print self.list_precio_iva_orig[self.i]
                self.i+=1
            f_i +=1

        self.header = self.tabla.horizontalHeader()  ##ajusta tabla al contenido
        self.header.setResizeMode(0, QtGui.QHeaderView.ResizeToContents)
        self.header.setResizeMode(1, QtGui.QHeaderView.ResizeToContents)
        self.header.setResizeMode(2, QtGui.QHeaderView.ResizeToContents)
        self.header.setResizeMode(3, QtGui.QHeaderView.ResizeToContents)
        self.tabla.setEditTriggers(QtGui.QAbstractItemView.NoEditTriggers)  # metodo para q no se puedan editar las columnas


   

'''
#datos a tomar en cuenta
#clicked() = click con el mouse
#returnpressed() = enter
#triggered() = para menu y barra de herramientas


        # conectamos los metodos de la barra con los eventos
        self.ui.barra_venta.triggered.connect(self.pagina_ventas)
        self.ui.barra_inventario.triggered.connect(self.pagina_inventario)
        self.ui.barra_usuarios.triggered.connect(self.pagina_usuarios)
        self.ui.barra_corte.triggered.connect(self.pagina_corte)
        self.ui.barra_salir.triggered.connect(self.salir)

        #conectamos los metodos de el menu con los eventos
        self.ui.menu_salir.triggered.connect(self.salir)
        self.ui.menu_info.triggered.connect(self.informacion)

########conectamos los botones con los eventos para ingresar, modificar y eliminar usuario###########
        self.connect(self.ui.ingresar_ingresar, SIGNAL("clicked()"),self.ingresar_usuario)          #
        self.connect(self.ui.ingresar_ingresar, SIGNAL("returnpressed()"),self.ingresar_usuario)    #
#---------------------------------------------------------------------------------------------------#
        self.connect(self.ui.modificar_buscar, SIGNAL("clicked()"),self.modificar_mostrar)          #
        self.connect(self.ui.modificar_buscar, SIGNAL("returnpressed()"), self.modificar_mostrar)   #
        self.connect(self.ui.modificar_modificar, SIGNAL("clicked()"), self.modificar_usuario)      #
        self.connect(self.ui.modificar_modificar, SIGNAL("returnpressed()"), self.modificar_usuario)#
#----------------qu-----------------------------------------------------------------------------------#
        self.connect(self.ui.eliminar_buscar, SIGNAL("clicked()"), self.eliminar_mostrar)           #
        self.connect(self.ui.eliminar_buscar, SIGNAL("returnpressed()"), self.eliminar_mostrar)     #
        self.connect(self.ui.eliminar_eliminar, SIGNAL("clicked()"), self.eliminar_eliminar)        #
        self.connect(self.ui.eliminar_eliminar, SIGNAL("returnpressed()"), self.eliminar_eliminar)  #
#####################################################################################################

########conectamos los botones con los eventos  para ingresar, modificar y eliminar productos########
        self.connect(self.ui.producto_ingresar, SIGNAL("clicked()"), self.ingresar_producto)        #
        self.connect(self.ui.producto_ingresar, SIGNAL("returnpressed()"), self.ingresar_producto)  #
#---------------------------------------------------------------------------------------------------#
        self.connect(self.ui.actualizar_buscar, SIGNAL("clicked()"), self.producto_mostrar)         #
        self.connect(self.ui.actualizar_buscar, SIGNAL("returnpressed()"), self.producto_mostrar)   #
        self.connect(self.ui.actualizar_actualizar, SIGNAL("clicked()"), self.producto_actualizar)  #
        self.connect(self.ui.actualizar_actualizar, SIGNAL("returnpressed()"), self.producto_actualizar)#
#---------------------------------------------------------------------------------------------------#
        self.connect(self.ui.eliminar_buscar_2, SIGNAL("clicked()"), self.eliminar_mostrar_producto)#
        self.connect(self.ui.eliminar_buscar_2, SIGNAL("returnpressed()"), self.eliminar_mostrar_producto)#
        self.connect(self.ui.eliminar_eliminar_2, SIGNAL("clicked()"), self.producto_eliminar)      #
        self.connect(self.ui.eliminar_eliminar_2, SIGNAL("returnpressed()"), self.producto_eliminar)#
####################################################################################################

#####################conectamos botones con los eventos de venta#####################################
        self.connect(self.ui.venta_buscar, SIGNAL("clicked()"), self.venta_mostrar)                 #
        self.connect(self.ui.venta_buscar, SIGNAL("returnpressed()"), self.venta_mostrar)           #
# --------------------------------------------------------------------------------------------------#
        self.ui.ventas_existencia.doubleClicked.connect(self.venta_click)                           #
        self.ui.ventas_final.doubleClicked.connect(self.quitar_producto)
          #
#  -------------------------------------------------------------------------------------------------#
        self.connect(self.ui.ventas_vender, SIGNAL("clicked()"), self.venta_final)                  #
        self.connect(self.ui.ventas_vender, SIGNAL("returnpressed()"), self.venta_final)
        self.connect(self.ui.ventas_cancelar, SIGNAL("clicked()"), self.cancelar_venta)
        self.connect(self.ui.ventas_cancelar, SIGNAL("returnpressed()"), self.cancelar_venta)
#####################################################################################################

#####################conectamos botones con los eventos de corte#####################################
        self.connect(self.ui.corte_generar, SIGNAL("clicked()"), self.archivo_mostrar)              #
        self.connect(self.ui.corte_generar, SIGNAL("returnpressed()"), self.archivo_mostrar)        #
#####################################################################################################

################lamado de paginas####################
    def pagina_ventas(self):                        #
        self.ui.paginas.setCurrentIndex(0)          #
        self.ui.ventas_existencia.clearContents()   #
        self.ui.ventas_existencia.setRowCount(0)    #
        self.ui.ventas_final.clearContents()        #
        self.ui.ventas_final.setRowCount(0)         #
#---------------------------------------------------#
    def pagina_inventario(self):                    #
        self.ui.paginas.setCurrentIndex(1)          #
        self.ui.actualizar_lista.clearContents()    #
        self.ui.actualizar_lista.setRowCount(0)     #
        self.ui.eliminar_lista_2.clearContents()    #
        self.ui.eliminar_lista_2.setRowCount(0)     #
#---------------------------------------------------#
    def pagina_usuarios(self):                      #
        self.ui.paginas.setCurrentIndex(2)          #
        self.ui.modificar_lista.clearContents()     #
        self.ui.modificar_lista.setRowCount(0)      #
        self.ui.eliminar_lista.clearContents()      #
        self.ui.eliminar_lista.setRowCount(0)       #
#---------------------------------------------------#
    def pagina_corte(self):                         #
        self.ui.paginas.setCurrentIndex(3)          #
        self.ui.corte_mostrar.clearContents()
        self.ui.corte_mostrar.setRowCount(0)
#####################################################


###########metodos para gestion de usuarios##########
    def ingresar_usuario(self):                     #
        root_usuarios.campo_vacio(self)             #
#---------------------------------------------------#
    def modificar_mostrar(self):                    #
        root_usuarios.buscar_usuarios(self)         #
        self.ui.modificar_modificar.setEnabled(True)#
#---------------------------------------------------#
    def modificar_usuario(self):                    #
        root_usuarios.modificar_vacio(self)         #
#---------------------------------------------------#
    def eliminar_mostrar(self):                     #
        root_usuarios.eliminar_mostrar(self)        #
        self.ui.eliminar_eliminar.setEnabled(True)  #
#---------------------------------------------------#
    def eliminar_eliminar(self):                    #
        root_usuarios.eliminar_usuario(self)        #
         #
#####################################################


##########metodos para gestion de productos##########
    def ingresar_producto(self):                    #
        root_almacen.campo_vacio(self)              #
#---------------------------------------------------#
    def producto_mostrar(self):                     #
        root_almacen.buscar_producto(self)          #
        self.ui.actualizar_actualizar.setEnabled(True)#
#---------------------------------------------------#
    def producto_actualizar(self):                  #
        root_almacen.alterar_producto(self)         #
#---------------------------------------------------#
    def eliminar_mostrar_producto(self):            #
        root_almacen.eliminar_mostrar(self)         #
        self.ui.eliminar_eliminar_2.setEnabled(True)#
#---------------------------------------------------#
    def producto_eliminar(self):                    #
        root_almacen.eliminar_producto(self)        #
        self.ui.eliminar_eliminar_2.setEnabled(False)#
#####################################################


#########metodos para venta de productos#############
    def venta_mostrar(self):                        #
        root_venta.ventas_mostrar(self)             #
#---------------------------------------------------#
    def venta_click(self):                          #
        self.ui.ventas_vender.setEnabled(True)      #
        self.ui.ventas_cancelar.setEnabled(True)    #
        root_venta.mandar_venta(self)               #
#---------------------------------------------------#
    def   quitar_producto(self):    #
        root_venta.eliminar_producto(self)
# ---------------------------------------------------#



    def venta_final(self):                          #
        self.ui.ventas_vender.setEnabled(False)     #
        self.ui.ventas_cancelar.setEnabled(False)
        root_venta.completar_venta(self, self.vendedor)            #
# --------------------------------------------------#
    def cancelar_venta(self):                       #
        self.ui.ventas_vender.setEnabled(False)
        self.ui.ventas_cancelar.setEnabled(False)  #
        root_venta.cancelar(self)                   #
#---------------------------------------------------#


#########metodos para generar corte##################
    def archivo_mostrar(self):                      #
        root_corte.mostrar_ventas(self, self.vendedor)   #
#---------------------------------------------------#
'
#metodo para informacion del programador
    def informacion(self):
        msg = QtGui.QMessageBox.about(self, "Acerca de",       Punto de venta
        realizado por Abuelazo
        correo mauro_ruiz2001@hotmail.com
                crostow.ewinkeiton@gmail.com
                                                 )
#metodo para salir
    def salir(self):
        self.close()
        ventana_1 = logueo.Logueo(self)
        ventana_1.show()



#metodo para centrar la ventana
    def centrado(self):
        screen = QtGui.QDesktopWidget().screenGeometry()
        size = self.geometry()
        self.move((screen.width() - size.width()) / 2, (screen.height() - size.height()) / 2)
'''
# codigo para lanzar la aplicacion
app = QtGui.QApplication(sys.argv)
MyWindow =  Buscador(None)

MyWindow.show()
app.exec_()
