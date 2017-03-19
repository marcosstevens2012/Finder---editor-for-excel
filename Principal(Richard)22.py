#!/usr/bin/env python
#-*- coding: utf-8 -*-
# -*- coding: cp1252 -*-

import sys
import sys,os
import sqlite3
from PyQt4.QtCore import *
from decimal import Decimal
from PyQt4 import QtGui, uic
import re
from PyQt4 import QtCore, QtGui, uic
from AcercaDe import acercade
from openpyxl import*



wb = Workbook()
# grab the active worksheet
ws = wb.active
import openpyxl   #Importamos libreria para abrir xls


##file_location = 'C:\Users\Richard\Dropbox\PROYECTO BUSCADOR/address.xlsx' #declaramos path o ruta del archivo a abrir

#se importa la pantalla ""from (carpeta) import (archivo)""

form_class = uic.loadUiType("Buscador.ui")[0]
###################################################################################

class Buscador(QtGui.QMainWindow, form_class):
    def __init__(self, parent=None):
        QtGui.QMainWindow.__init__(self, parent)
        self.setupUi(self)
        #self.l_cod_b.setValidator(QtGui.QIntValidator())
        self.l_carga.hide() == True ##oculta el label carga
        self.btn_cargar.setEnabled(False) 
        self.array_2_table()
        self.btn_cargar.clicked.connect(self.btn_cargar_clicked)##metodo carga archivo
        self.btn_AcercaDe.clicked.connect(self.btn_AcercaDe_clicked)
        self.l_porc_b.textChanged.connect(self.l_porc_b_Changed)
        self.tabla.clicked.connect(self.mostrarDatos)
        self.l_cod_b.setEnabled(False)
        self.l_cod.setEnabled(False)
        self.l_des.setEnabled(False)
       

        
        self.barra.setValue(0)


######################################################################################################################


    def array_2_table(self):
        self.tabla.setRowCount(18000)
        QtGui.QMessageBox.information( self,'Informacion', '''Por favor defina porcentaje de ganancia antes de cargar el archivo''', QtGui.QMessageBox.Ok)



######################################################################################################################
    def btn_Buscar_clicked(self):
	 	srows = self.tabla.rowCount()
	 	for row in xrange(18, srows):
		 	texto = str(self.tabla.item(row, 2))
		 	
     		print texto

              
#########################################################################################################
    def mostrarDatos(self):#muestra los datos en los labels
        self.tabla.setSelectionBehavior(self.tabla.SelectRows)
        self.descripcion = self.tabla.item(0, 1).text()
        #self.nomGrupo = self.tableGrupos.item(i, 2).text()
        #self.estado = self.tableGrupos.item(i, 1).text()
        self.l_descripcion.setText(str(self.descripcion))


    #if (l_porc_b.isModified):
    def l_porc_b_Changed(self): 
  		self.btn_cargar.setEnabled(True)


    def btn_AcercaDe_clicked(self):
        from AcercaDe import acercade
     	ventana_2 = acercade(self)
        #se muestra la ventana
     	ventana_2.show()


#   #######################################################################
    def btn_cargar_clicked(self):#abre el explorador de archivos
    	self.filePath = QtGui.QFileDialog.getOpenFileName(self,'Cargar Archivo',"~/Desktop/",'*.xlsx')
        #doc = openpyxl.load_workbook(str(self.filePath))
        #hoja = doc.get_sheet_by_name('Sheet1')  # abre el archivo #se debe poner en el metodo a usar
        self.l_carga.show()#muestra el label carga
        self.l_carga.setStyleSheet("QLabel#l_carga {color: green}")#cambia de color el label carga
        self.l_cod_b.setEnabled(True)
        self.l_cod.setEnabled(True)
        self.l_des.setEnabled(True)
        
        print(str(self.filePath))
        doc = openpyxl.load_workbook(str(self.filePath))
        hoja = doc.get_sheet_by_name('Hoja1')


        self.barra.setValue(10)

        seleccionCodigo = hoja['A18':'A17833'] 
      	seleccionDesc = hoja['B18':'B17833']
      	seleccionScanner = hoja['F18':'F17833']
      	seleccionIVA = hoja['K18':'K17833']

  	

      	porcentaje = self.l_porc_b.text()
      	porcentaje2 = (float(porcentaje) / 100)
      	p = 0.0+porcentaje2
  	
  
       	f_i = 0
       	for filas in seleccionCodigo:
    	 for celda in filas:
           self.barra.setValue(30)
      	   self.tabla.setItem(f_i,3, QtGui.QTableWidgetItem(str(celda.value)))
         f_i +=1

      	f_i = 0
      	for filas in seleccionDesc:
    	 	for celda in filas:
      		 	self.tabla.setItem(f_i,1, QtGui.QTableWidgetItem(celda.value))
                self.barra.setValue(40)
    	 	f_i +=1

     
      	f_i = 0
      	for filas in seleccionScanner:
         for celda in filas:
           self.tabla.setItem(f_i,0, QtGui.QTableWidgetItem(str(celda.value)))
           self.barra.setValue(70)
         f_i +=1

      	f_i = 0
      	p1 = 0
      	for filas in seleccionIVA:
         for celda in filas:
            p1 = celda.value*porcentaje2
            self.tabla.setItem(f_i,2, QtGui.QTableWidgetItem(str(p1+celda.value)))
            self.barra.setValue(100)
         f_i +=1

        self.header = self.tabla.horizontalHeader()  ##ajusta tabla al contenido
        self.header.setResizeMode(0, QtGui.QHeaderView.ResizeToContents)
        self.header.setResizeMode(1, QtGui.QHeaderView.ResizeToContents)
        self.header.setResizeMode(2, QtGui.QHeaderView.ResizeToContents)
        self.header.setResizeMode(3, QtGui.QHeaderView.ResizeToContents)
        self.tabla.setEditTriggers(QtGui.QAbstractItemView.NoEditTriggers)  # metodo para q no se puedan editar las columnas


        
 
'''
#datos a tomar en cuenta
#clicked() = click con el mouse
#returnpressed() = enter
#triggered() = para menu y barra de herramientas


        # conectamos los metodos de la barra con los eventos
        self.ui.barra_venta.triggered.connect(self.pagina_ventas)
        self.ui.barra_inventario.triggered.connect(self.pagina_inventario)
        self.ui.barra_usuarios.triggered.connect(self.pagina_usuarios)
        self.ui.barra_corte.triggered.connect(self.pagina_corte)
        self.ui.barra_salir.triggered.connect(self.salir)

        #conectamos los metodos de el menu con los eventos
        self.ui.menu_salir.triggered.connect(self.salir)
        self.ui.menu_info.triggered.connect(self.informacion)

########conectamos los botones con los eventos para ingresar, modificar y eliminar usuario###########
        self.connect(self.ui.ingresar_ingresar, SIGNAL("clicked()"),self.ingresar_usuario)          #
        self.connect(self.ui.ingresar_ingresar, SIGNAL("returnpressed()"),self.ingresar_usuario)    #
#---------------------------------------------------------------------------------------------------#
        self.connect(self.ui.modificar_buscar, SIGNAL("clicked()"),self.modificar_mostrar)          #
        self.connect(self.ui.modificar_buscar, SIGNAL("returnpressed()"), self.modificar_mostrar)   #
        self.connect(self.ui.modificar_modificar, SIGNAL("clicked()"), self.modificar_usuario)      #
        self.connect(self.ui.modificar_modificar, SIGNAL("returnpressed()"), self.modificar_usuario)#
#----------------qu-----------------------------------------------------------------------------------#
        self.connect(self.ui.eliminar_buscar, SIGNAL("clicked()"), self.eliminar_mostrar)           #
        self.connect(self.ui.eliminar_buscar, SIGNAL("returnpressed()"), self.eliminar_mostrar)     #
        self.connect(self.ui.eliminar_eliminar, SIGNAL("clicked()"), self.eliminar_eliminar)        #
        self.connect(self.ui.eliminar_eliminar, SIGNAL("returnpressed()"), self.eliminar_eliminar)  #
#####################################################################################################

########conectamos los botones con los eventos  para ingresar, modificar y eliminar productos########
        self.connect(self.ui.producto_ingresar, SIGNAL("clicked()"), self.ingresar_producto)        #
        self.connect(self.ui.producto_ingresar, SIGNAL("returnpressed()"), self.ingresar_producto)  #
#---------------------------------------------------------------------------------------------------#
        self.connect(self.ui.actualizar_buscar, SIGNAL("clicked()"), self.producto_mostrar)         #
        self.connect(self.ui.actualizar_buscar, SIGNAL("returnpressed()"), self.producto_mostrar)   #
        self.connect(self.ui.actualizar_actualizar, SIGNAL("clicked()"), self.producto_actualizar)  #
        self.connect(self.ui.actualizar_actualizar, SIGNAL("returnpressed()"), self.producto_actualizar)#
#---------------------------------------------------------------------------------------------------#
        self.connect(self.ui.eliminar_buscar_2, SIGNAL("clicked()"), self.eliminar_mostrar_producto)#
        self.connect(self.ui.eliminar_buscar_2, SIGNAL("returnpressed()"), self.eliminar_mostrar_producto)#
        self.connect(self.ui.eliminar_eliminar_2, SIGNAL("clicked()"), self.producto_eliminar)      #
        self.connect(self.ui.eliminar_eliminar_2, SIGNAL("returnpressed()"), self.producto_eliminar)#
####################################################################################################

#####################conectamos botones con los eventos de venta#####################################
        self.connect(self.ui.venta_buscar, SIGNAL("clicked()"), self.venta_mostrar)                 #
        self.connect(self.ui.venta_buscar, SIGNAL("returnpressed()"), self.venta_mostrar)           #
# --------------------------------------------------------------------------------------------------#
        self.ui.ventas_existencia.doubleClicked.connect(self.venta_click)                           #
        self.ui.ventas_final.doubleClicked.connect(self.quitar_producto)
          #
#  -------------------------------------------------------------------------------------------------#
        self.connect(self.ui.ventas_vender, SIGNAL("clicked()"), self.venta_final)                  #
        self.connect(self.ui.ventas_vender, SIGNAL("returnpressed()"), self.venta_final)
        self.connect(self.ui.ventas_cancelar, SIGNAL("clicked()"), self.cancelar_venta)
        self.connect(self.ui.ventas_cancelar, SIGNAL("returnpressed()"), self.cancelar_venta)
#####################################################################################################

#####################conectamos botones con los eventos de corte#####################################
        self.connect(self.ui.corte_generar, SIGNAL("clicked()"), self.archivo_mostrar)              #
        self.connect(self.ui.corte_generar, SIGNAL("returnpressed()"), self.archivo_mostrar)        #
#####################################################################################################

################lamado de paginas####################
    def pagina_ventas(self):                        #
        self.ui.paginas.setCurrentIndex(0)          #
        self.ui.ventas_existencia.clearContents()   #
        self.ui.ventas_existencia.setRowCount(0)    #
        self.ui.ventas_final.clearContents()        #
        self.ui.ventas_final.setRowCount(0)         #
#---------------------------------------------------#
    def pagina_inventario(self):                    #
        self.ui.paginas.setCurrentIndex(1)          #
        self.ui.actualizar_lista.clearContents()    #
        self.ui.actualizar_lista.setRowCount(0)     #
        self.ui.eliminar_lista_2.clearContents()    #
        self.ui.eliminar_lista_2.setRowCount(0)     #
#---------------------------------------------------#
    def pagina_usuarios(self):                      #
        self.ui.paginas.setCurrentIndex(2)          #
        self.ui.modificar_lista.clearContents()     #
        self.ui.modificar_lista.setRowCount(0)      #
        self.ui.eliminar_lista.clearContents()      #
        self.ui.eliminar_lista.setRowCount(0)       #
#---------------------------------------------------#
    def pagina_corte(self):                         #
        self.ui.paginas.setCurrentIndex(3)          #
        self.ui.corte_mostrar.clearContents()
        self.ui.corte_mostrar.setRowCount(0)
#####################################################


###########metodos para gestion de usuarios##########
    def ingresar_usuario(self):                     #
        root_usuarios.campo_vacio(self)             #
#---------------------------------------------------#
    def modificar_mostrar(self):                    #
        root_usuarios.buscar_usuarios(self)         #
        self.ui.modificar_modificar.setEnabled(True)#
#---------------------------------------------------#
    def modificar_usuario(self):                    #
        root_usuarios.modificar_vacio(self)         #
#---------------------------------------------------#
    def eliminar_mostrar(self):                     #
        root_usuarios.eliminar_mostrar(self)        #
        self.ui.eliminar_eliminar.setEnabled(True)  #
#---------------------------------------------------#
    def eliminar_eliminar(self):                    #
        root_usuarios.eliminar_usuario(self)        #
         #
#####################################################


##########metodos para gestion de productos##########
    def ingresar_producto(self):                    #
        root_almacen.campo_vacio(self)              #
#---------------------------------------------------#
    def producto_mostrar(self):                     #
        root_almacen.buscar_producto(self)          #
        self.ui.actualizar_actualizar.setEnabled(True)#
#---------------------------------------------------#
    def producto_actualizar(self):                  #
        root_almacen.alterar_producto(self)         #
#---------------------------------------------------#
    def eliminar_mostrar_producto(self):            #
        root_almacen.eliminar_mostrar(self)         #
        self.ui.eliminar_eliminar_2.setEnabled(True)#
#---------------------------------------------------#
    def producto_eliminar(self):                    #
        root_almacen.eliminar_producto(self)        #
        self.ui.eliminar_eliminar_2.setEnabled(False)#
#####################################################


#########metodos para venta de productos#############
    def venta_mostrar(self):                        #
        root_venta.ventas_mostrar(self)             #
#---------------------------------------------------#
    def venta_click(self):                          #
        self.ui.ventas_vender.setEnabled(True)      #
        self.ui.ventas_cancelar.setEnabled(True)    #
        root_venta.mandar_venta(self)               #
#---------------------------------------------------#
    def   quitar_producto(self):    #
        root_venta.eliminar_producto(self)
# ---------------------------------------------------#



    def venta_final(self):                          #
        self.ui.ventas_vender.setEnabled(False)     #
        self.ui.ventas_cancelar.setEnabled(False)
        root_venta.completar_venta(self, self.vendedor)            #
# --------------------------------------------------#
    def cancelar_venta(self):                       #
        self.ui.ventas_vender.setEnabled(False)
        self.ui.ventas_cancelar.setEnabled(False)  #
        root_venta.cancelar(self)                   #
#---------------------------------------------------#


#########metodos para generar corte##################
    def archivo_mostrar(self):                      #
        root_corte.mostrar_ventas(self, self.vendedor)   #
#---------------------------------------------------#
'
#metodo para informacion del programador
    def informacion(self):
        msg = QtGui.QMessageBox.about(self, "Acerca de",       Punto de venta
        realizado por Abuelazo
        correo mauro_ruiz2001@hotmail.com
                crostow.ewinkeiton@gmail.com
                                                 )
#metodo para salir
    def salir(self):
        self.close()
        ventana_1 = logueo.Logueo(self)
        ventana_1.show()



#metodo para centrar la ventana
    def centrado(self):
        screen = QtGui.QDesktopWidget().screenGeometry()
        size = self.geometry()
        self.move((screen.width() - size.width()) / 2, (screen.height() - size.height()) / 2)
'''
# codigo para lanzar la aplicacion
app = QtGui.QApplication(sys.argv)
MyWindow =  Buscador(None)

MyWindow.show()
app.exec_()
