#!/usr/bin/env python
#-*- coding: utf-8 -*-
# -*- coding: cp1252 -*-
import sys
reload(sys)
sys.setdefaultencoding('utf8')
from PyQt4 import QtGui, uic
from AcercaDe import acercade
from openpyxl import*#Importamos libreria para abrir xls
import openpyxl

wb = Workbook()
# grab the active worksheet
ws = wb.active

##file_location = 'C:\Users\Richard\Dropbox\PROYECTO BUSCADOR/address.xlsx' #declaramos path o ruta del archivo a abrir

#se importa la pantalla ""from (carpeta) import (archivo)""

form_class = uic.loadUiType("Buscador.ui")[0]
###################################################################################

class Buscador(QtGui.QMainWindow, form_class):
    def __init__(self, parent=None):
        QtGui.QMainWindow.__init__(self, parent)
        self.setupUi(self)
        #self.l_cod_b.setValidator(QtGui.QIntValidator())
        self.l_carga.hide() == True ##oculta el label carga
        self.btn_cargar.setEnabled(False) 
        self.array_2_table()
        QtGui.QMessageBox.information( self,'Informacion', '''Por favor defina porcentaje de ganancia antes de cargar el archivo''', QtGui.QMessageBox.Ok)
        self.l_des.textEdited.connect(lambda text: self.l_des.setText(text.toUpper()))#pone en mayusculas el lineEdit

        self.l_cod.setValidator(QtGui.QIntValidator())
        self.l_porc_b.setValidator(QtGui.QIntValidator())

        self.btn_cargar.clicked.connect(self.btn_cargar_clicked)##metodo carga archivo
        self.btn_AcercaDe.clicked.connect(self.btn_AcercaDe_clicked)
        self.btn_aplicar.clicked.connect(self.btn_aplicar_clicked)
        self.l_porc_b.textChanged.connect(self.l_porc_b_Changed)
        self.tabla.clicked.connect(self.mostrarDatos)
        self.l_cod.textChanged.connect(self.buscar_codigo)
        self.l_des.textChanged.connect(self.buscar_codigo_descripcion)
        self.l_cod_b.textChanged.connect(self.buscar_codigo_scanner)
        self.l_cod_b.setEnabled(False)
        self.l_cod.setEnabled(False)
        self.l_des.setEnabled(False)
        self.l_porc_b.setFocus()
        self.list_precio_iva_orig=[]#lista con los precios son iva original
        self.list_precio_iva_new=[]#lista con los precios iva con ganancia
        self.list_codigo=[]#carga los codigo
        self.list_codigo_scanner=[]
        self.list_descripcion=[]
        self.btn_aplicar.setEnabled(False)

######################################################################################################################


    def array_2_table(self):
        self.tabla.setRowCount(1000)



######################################################################################################################
    def buscar_codigo(self):
        self.l_cod_b.clear()
        self.l_des.clear()
        while self.tabla.rowCount() > 0:  # limpia la tabla de pacientes
            self.tabla.removeRow(0)
        self.tabla.setRowCount(200)
        self.codigo_b = str(self.l_cod.text())
        #Se agregan los elementos al QListWidget
        self.lon=len(self.l_cod.text())
        i = 0
        y=-1
        for x in self.list_codigo:
            y+=1
            if (self.codigo_b)== str(x)[0:self.lon]:
                self.tabla.setItem(i,0, QtGui.QTableWidgetItem(self.list_codigo_scanner[y]))
                self.tabla.setItem(i, 1, QtGui.QTableWidgetItem(self.list_descripcion[y]))
                self.tabla.setItem(i,2, QtGui.QTableWidgetItem("{0:.2f}".format(self.list_precio_iva_new[y])))#se redondeo el valor a dos decimales
                self.tabla.setItem(i, 3, QtGui.QTableWidgetItem(str(self.list_codigo[y])))
                self.tabla.setEditTriggers(QtGui.QAbstractItemView.NoEditTriggers)#metodo para q no se puedan editar las columnas
                self.tabla.setSelectionBehavior(self.tabla.SelectRows)
                i+=1
#########################################################################################################

    def buscar_codigo_descripcion(self):
        self.l_cod_b.clear()
        self.l_cod.clear()
        while self.tabla.rowCount() > 0:  # limpia la tabla de pacientes
            self.tabla.removeRow(0)
        self.tabla.setRowCount(200)
        self.des = str(self.l_des.text())
        # Se agregan los elementos al QListWidget
        self.lon = len(self.l_des.text())
        i = 0
        y = -1
        for x in self.list_descripcion:
            y += 1
            if (self.des) == str(x)[0:self.lon]:
                self.tabla.setItem(i, 0, QtGui.QTableWidgetItem(self.list_codigo_scanner[y]))
                self.tabla.setItem(i, 1, QtGui.QTableWidgetItem(self.list_descripcion[y]))
                self.tabla.setItem(i, 2, QtGui.QTableWidgetItem("{0:.2f}".format(self.list_precio_iva_new[y])))  # se redondeo el valor a dos decimales
                self.tabla.setItem(i, 3, QtGui.QTableWidgetItem(str(self.list_codigo[y])))
                self.tabla.setEditTriggers(QtGui.QAbstractItemView.NoEditTriggers)  # metodo para q no se puedan editar las columnas
                self.tabla.setSelectionBehavior(self.tabla.SelectRows)
                i += 1
###########################################################################################################################

    def buscar_codigo_scanner(self):
        self.l_cod.clear()
        self.l_des.clear()
        while self.tabla.rowCount() > 0:  # limpia la tabla de pacientes
            self.tabla.removeRow(0)
        self.tabla.setRowCount(200)
        self.codigo_b = str(self.l_cod_b.text())
        # Se agregan los elementos al QListWidget
        self.lon = len(self.l_cod_b.text())
        i = 0
        y = -1
        for x in self.list_codigo_scanner:
            y += 1
            if (self.codigo_b) == str(x)[0:self.lon]:
                self.tabla.setItem(i, 0, QtGui.QTableWidgetItem(self.list_codigo_scanner[y]))
                self.tabla.setItem(i, 1, QtGui.QTableWidgetItem(self.list_descripcion[y]))
                self.tabla.setItem(i, 2, QtGui.QTableWidgetItem("{0:.2f}".format(self.list_precio_iva_new[y])))  # se redondeo el valor a dos decimales
                self.tabla.setItem(i, 3, QtGui.QTableWidgetItem(str(self.list_codigo[y])))
                self.tabla.setEditTriggers(QtGui.QAbstractItemView.NoEditTriggers)  # metodo para q no se puedan editar las columnas
                self.tabla.setSelectionBehavior(self.tabla.SelectRows)
                i += 1

#########################################################################################################
    def mostrarDatos(self):#muestra los datos en los labels
        self.tabla.setSelectionBehavior(self.tabla.SelectRows)
        rows = self.tabla.selectionModel().selectedRows()
        index = []
        for i in rows:
            index.append(i.row())
        for i in index:
            self.descripcion = self.tabla.item(i, 1).text()
            self.precio=self.tabla.item(i, 2).text()
            #self.nomGrupo = self.tableGrupos.item(i, 2).text()
            #self.estado = self.tableGrupos.item(i, 1).text()
            self.l_descripcion.setText(str(self.descripcion))
            self.l_precio_iva.setText(str(self.precio))
            self.precio_orig_iva=("{0:.2f}".format(self.list_precio_iva_orig[i]))
            self.l_ganancia.setText(str(float(self.precio)-float(self.precio_orig_iva)))


    #if (l_porc_b.isModified):
    def l_porc_b_Changed(self): 
  		self.btn_cargar.setEnabled(True)


    def btn_AcercaDe_clicked(self):
        from AcercaDe import acercade
     	ventana_2 = acercade(self)
        #se muestra la ventana
     	ventana_2.show()


#   #######################################################################
    def btn_cargar_clicked(self):  # abre el explorador de archivos
        self.filePath = QtGui.QFileDialog.getOpenFileName(self, 'Cargar Archivo', "~/Desktop/", '*.xlsx')
        # doc = openpyxl.load_workbook(str(self.filePath))
        # hoja = doc.get_sheet_by_name('Sheet1')  # abre el archivo #se debe poner en el metodo a usar

        self.l_cod_b.setEnabled(True)
        self.l_cod.setEnabled(True)
        self.l_des.setEnabled(True)
        # print(str(self.filePath))
        self.barra.setValue(0)
        self.doc = openpyxl.load_workbook(str(self.filePath))
        self.hoja = self.doc.get_sheet_by_name('Hoja1')

        seleccionCodigo = self.hoja['A18':'A17833']
        seleccionDesc = self.hoja['B18':'B17833']
        seleccionScanner = self.hoja['F18':'F17833']
        seleccionIVA = self.hoja['K18':'K17833']

        self.porcentaje = self.l_porc_b.text()
        self.porcentaje2 = (float(self.porcentaje) / 100)
        p = 0.0 + self.porcentaje2

        f_i = 0
        self.barra.setValue(15)
        for filas in seleccionCodigo:
            for celda in filas:
                self.tabla.setItem(f_i, 3, QtGui.QTableWidgetItem(str(celda.value)))
                self.list_codigo.append(celda.value)
            f_i += 1
        self.barra.setValue(30)
        f_i = 0
        for filas in seleccionDesc:
            for celda in filas:
                self.tabla.setItem(f_i, 1, QtGui.QTableWidgetItem(celda.value))
                self.list_descripcion.append(celda.value)
            f_i += 1

        f_i = 0
        self.barra.setValue(50)
        for filas in seleccionScanner:
            for celda in filas:
                self.tabla.setItem(f_i, 0, QtGui.QTableWidgetItem(str(celda.value)))
                self.list_codigo_scanner.append(celda.value)
            f_i += 1

        f_i = 0
        self.p1 = 0
        self.i = 0
        self.barra.setValue(70)
        for filas in seleccionIVA:
            for celda in filas:
                self.p1 = celda.value * self.porcentaje2
                self.tabla.setItem(f_i, 2, QtGui.QTableWidgetItem("{0:.2f}".format(self.p1 + celda.value)))
                self.list_precio_iva_new.append(self.p1 + celda.value)  # agrega datos a la lista en la ulrima posicion
                self.list_precio_iva_orig.append(celda.value)
                # print self.list_precio_iva_orig[self.i]
                self.i += 1
            f_i += 1
        self.barra.setValue(100)
        self.l_carga.show()  # muestra el label carga
        self.l_carga.setStyleSheet("QLabel#l_carga {color: green}")  # cambia de color el label carga
        self.header = self.tabla.horizontalHeader()  ##ajusta tabla al contenido
        self.header.setResizeMode(0, QtGui.QHeaderView.ResizeToContents)
        self.header.setResizeMode(1, QtGui.QHeaderView.ResizeToContents)
        self.header.setResizeMode(2, QtGui.QHeaderView.ResizeToContents)
        self.header.setResizeMode(3, QtGui.QHeaderView.ResizeToContents)
        self.tabla.setEditTriggers(QtGui.QAbstractItemView.NoEditTriggers)  # metodo para q no se puedan editar las columnas
        self.btn_aplicar.setEnabled(True)


    def btn_aplicar_clicked(self):#abre el explorador de archivo
        self.barra.setValue(00)
        self.barra.setValue(10)

        seleccionCodigo = self.hoja['A18':'A17833']
        seleccionDesc = self.hoja['B18':'B17833']
        seleccionScanner = self.hoja['F18':'F17833']
        seleccionIVA = self.hoja['K18':'K17833']


        self.porcentaje = self.l_porc_b.text()
        self.porcentaje2 = (float(self.porcentaje) / 100)
        p = 0.0+self.porcentaje2
        self.barra.setValue(20)
        f_i = 0

        for filas in seleccionCodigo:
            for celda in filas:
                self.tabla.setItem(f_i,3, QtGui.QTableWidgetItem(str(celda.value)))
                self.list_codigo.append(celda.value)
            f_i +=1
        self.barra.setValue(50)
        f_i = 0
        for filas in seleccionDesc:
                for celda in filas:
                        self.tabla.setItem(f_i,1, QtGui.QTableWidgetItem(celda.value))
                        self.list_descripcion.append(celda.value)
                f_i +=1

        self.barra.setValue(70)
        f_i = 0
        for filas in seleccionScanner:
                for celda in filas:
                    self.tabla.setItem(f_i,0, QtGui.QTableWidgetItem(str(celda.value)))
                    self.list_codigo_scanner.append(celda.value)
                f_i +=1

        f_i = 0
        p1 = 0
        self.i=0
        for filas in seleccionIVA:
            for celda in filas:
                self.p1 = celda.value*self.porcentaje2
                self.tabla.setItem(f_i,2, QtGui.QTableWidgetItem("{0:.2f}".format(self.p1+celda.value)))
                self.list_precio_iva_new.append(self.p1+celda.value)#agrega datos a la lista en la ulrima posicion
                #print self.list_precio_iva_orig[self.i]
                self.i+=1
            f_i +=1
        self.barra.setValue(100)
        self.l_carga.show()  # muestra el label carga
        self.l_carga.setStyleSheet("QLabel#l_carga {color: green}")  # cambia de color el label carga
        self.header = self.tabla.horizontalHeader()  ##ajusta tabla al contenido
        self.header.setResizeMode(0, QtGui.QHeaderView.ResizeToContents)
        self.header.setResizeMode(1, QtGui.QHeaderView.ResizeToContents)
        self.header.setResizeMode(2, QtGui.QHeaderView.ResizeToContents)
        self.header.setResizeMode(3, QtGui.QHeaderView.ResizeToContents)
        self.tabla.setEditTriggers(QtGui.QAbstractItemView.NoEditTriggers)  # metodo para q no se puedan editar las columnas
        self.btn_aplicar.setEnabled(True)
# codigo para lanzar la aplicacion
app = QtGui.QApplication(sys.argv)
MyWindow =  Buscador(None)

MyWindow.show()
app.exec_()
